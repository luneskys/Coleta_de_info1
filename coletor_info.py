import os
import socket
import getpass
import platform
import psutil
from openpyxl import Workbook, load_workbook
import cpuinfo
import winreg
import tkinter as tk
from tkinter import messagebox
import logging
import multiprocessing
import sys
import subprocess

# Configuração de logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def coletar_informacoes():
    logging.debug("Iniciando coleta de informações...")
    informacoes = {}
    try:
        informacoes['Nome do Computador'] = os.getenv('COMPUTERNAME')
        informacoes['Endereço IP'] = socket.gethostbyname(socket.gethostname())
        informacoes['Nome do Usuário'] = getpass.getuser()
        
        # Nome do Domínio
        tentativas = 0
        while tentativas < 2:
            try:
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SYSTEM\CurrentControlSet\Services\Tcpip\Parameters")
                domain, _ = winreg.QueryValueEx(key, "Domain")
                if domain:
                    informacoes['Domínio'] = domain
                    break
                else:
                    tentativas += 1
            except FileNotFoundError:
                logging.error("Chave de registro não encontrada para domínio.")
                tentativas += 1
            except Exception as e:
                logging.error("Erro ao obter domínio: %s", e)
                tentativas += 1
        if tentativas == 2:
            informacoes['Domínio'] = "Desconhecido"
        
        # Sistema Operacional e Versão
        so_info = platform.uname()
        try:
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows NT\CurrentVersion")
            edition, _ = winreg.QueryValueEx(key, "EditionID")
            informacoes['Sistema Operacional'] = f"{so_info.system} {so_info.release} {edition}"
        except Exception as e:
            logging.error("Erro ao obter informações do sistema operacional: %s", e)
            informacoes['Sistema Operacional'] = f"{so_info.system} {so_info.release} Desconhecida"
        
        # Processador
        cpu_info = cpuinfo.get_cpu_info()
        # Ajustar o formato do processador
        processador = cpu_info['brand_raw']
        processador = processador.split(' ')[-3] + ' ' + ' '.join(processador.split(' ')[-2:])
        informacoes['Processador'] = processador
        
        # Memória Total
        informacoes['Memória Total (GB)'] = round(psutil.virtual_memory().total / (1024 ** 3))  # Arredondar para cima
        
        # Modelo do Computador
        try:
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"HARDWARE\DESCRIPTION\System\BIOS")
            model, _ = winreg.QueryValueEx(key, "SystemProductName")
            informacoes['Modelo do Computador'] = model
        except Exception as e:
            logging.error("Erro ao obter modelo do computador: %s", e)
            informacoes['Modelo do Computador'] = "Desconhecido"
        
        # Número de Série
        try:
            serial_number = subprocess.check_output(['powershell', 'Get-WmiObject -Class Win32_BIOS | Select-Object -Property SerialNumber']).decode().split('\n')[3].strip()
            informacoes['Número de Série'] = serial_number
        except Exception as e:
            logging.error("Erro ao obter número de série: %s", e)
            informacoes['Número de Série'] = "Desconhecido"
        
        logging.debug("Informações coletadas: %s", informacoes)
    except Exception as e:
        logging.error("Erro ao coletar informações: %s", e)
    
    return informacoes

def atualizar_planilha(informacoes):
    logging.debug("Iniciando atualização da planilha...")
    # Obter o diretório onde o executável está localizado
    if getattr(sys, 'frozen', False):
        diretorio_atual = os.path.dirname(sys.executable)
    else:
        diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    arquivo = os.path.join(diretorio_atual, 'informacoes_computadores.xlsx')

    try:
        if os.path.exists(arquivo):
            workbook = load_workbook(arquivo)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Nome do Computador', 'Nome do Usuário', 'Domínio', 'Número de Série', 'Modelo do Computador', 'Processador', 'Memória Total (GB)', 'Sistema Operacional', 'Endereço IP'])

        sheet.append([informacoes['Nome do Computador'], informacoes['Nome do Usuário'], informacoes['Domínio'], informacoes['Número de Série'], informacoes['Modelo do Computador'], informacoes['Processador'], informacoes['Memória Total (GB)'], informacoes['Sistema Operacional'], informacoes['Endereço IP']])
        workbook.save(arquivo)
        logging.debug("Planilha atualizada com sucesso: %s", arquivo)
        mostrar_alerta(f"Planilha atualizada com sucesso: {arquivo}")
    except PermissionError:
        logging.error("Erro de permissão ao tentar salvar o arquivo: %s", arquivo)
        mostrar_alerta(f"Erro de permissão ao tentar salvar o arquivo: {arquivo}. Verifique se o arquivo está aberto ou se você tem permissões de escrita.")
    except Exception as e:
        logging.error("Ocorreu um erro ao tentar salvar o arquivo: %s", e)
        mostrar_alerta(f"Ocorreu um erro ao tentar salvar o arquivo: {e}")

def mostrar_alerta(mensagem):
    root = tk.Tk()
    root.withdraw()  # Ocultar a janela principal
    messagebox.showinfo("Aviso", mensagem)
    root.after(2000, root.destroy)  # Fechar a janela após 2 segundos

def main():
    informacoes = coletar_informacoes()
    atualizar_planilha(informacoes)
    input("Pressione Enter para sair...")

if __name__ == "__main__":
    # PyInstaller fix
    multiprocessing.freeze_support()
    main()
